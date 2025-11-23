import numpy as np
import cvxpy as cp
import time
import matplotlib.pyplot as plt
import openpyxl

book = openpyxl.load_workbook(filename='SAGA.xlsx')

sheet = book.worksheets[0]
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Sheet1"

np.random.seed(42)
print("Setting up Mean-Downside-Risk + L1 (Prox-SAGA) problem...")


R = np.loadtxt("return.csv", delimiter=",")
T, N = R.shape
mu = R.mean(axis=0).reshape(-1, 1)

print(f"Data loaded: {T} observations, {N} assets.")


params = {}
params["lambda"] = 1.0
params["b"] = 0.0
params["alpha_L1"] = 5e-4

params["R"] = R
params["mu"] = mu
params["T"] = T
params["N"] = N

params["max_epochs"] = 100
params["tol"] = 1e-8


L_max = 2 * params["lambda"] * np.max(np.sum(R * R, axis=1))
params["step_size"] = 1.0 / (3 * L_max)

print(f"L1 penalty alpha_L1: {params['alpha_L1']:e}")
print(f"SAGA Max Lipschitz L_max: {L_max:e}")
print(f"SAGA Step Size t: {params['step_size']:e}")


def calc_objective(w, p):

    Rw = p["R"] @ w
    loss = np.maximum(0, p["b"] - Rw)**2

    term1 = -(w.T @ p["mu"]).item()
    term2 = (p["lambda"] / p["T"]) * np.sum(loss)
    term3 = p["alpha_L1"] * np.sum(np.abs(w))

    return term1 + term2 + term3


def prox_operator(z, t, p):

    return np.sign(z) * np.maximum(np.abs(z) - t * p["alpha_L1"], 0)


def calc_f_i_grad(w, i, p):

    r_i = p["R"][i].reshape(-1, 1)
    loss = p["b"] - (r_i.T @ w).item()

    if loss > 0:
        return 2 * p["lambda"] * loss * (-r_i)
    return np.zeros_like(w)



w = np.zeros((N, 1))
t = params["step_size"]

grad_table = np.zeros((T, N))
obj_history = []

print(f"Initializing SAGA gradient table (T={T})...")

for i in range(T):
    grad_table[i] = calc_f_i_grad(w, i, params).reshape(-1)

grad_avg = grad_table.mean(axis=0).reshape(-1, 1)

print("Initialization complete.")

cnt = 0

start_time = time.time()

for epoch in range(1, params["max_epochs"] + 1):

    w_old_epoch = w.copy()
    obj_history.append(calc_objective(w, params))


    grad_h = -params["mu"]
    idx_perm = np.random.permutation(T)

    for it in idx_perm:
        grad_f_old = grad_table[it].reshape(-1, 1)
        grad_f_new = calc_f_i_grad(w, it, params)

        G_k = grad_h + (grad_f_new - grad_f_old) + grad_avg

        w = prox_operator(w - t * G_k, t, params)

        grad_avg += (grad_f_new - grad_f_old) / T
        grad_table[it] = grad_f_new.reshape(-1)
    if epoch > 1:
        print(f"Epoch: {epoch-1:4d}, Obj: {obj_history[-2]:e}, Step Norm: {np.linalg.norm(w - w_old_epoch):e}")
    cnt = cnt + 1
    ws.cell(column=1,row=cnt).value = calc_objective(w, params)
    # Stopping criterion
    if np.linalg.norm(w - w_old_epoch) < params["tol"]* (1 + np.linalg.norm(w_old_epoch)):
        print("Convergence reached! Step size tolerance met.")
        break

runtime = time.time() - start_time
print("...SAGA finished.")

wb.save("SAGA.xlsx")

w_opt = w
final_obj_saga = obj_history[-1]

print("\n--- Prox-SAGA Optimization Results ---")
print(f"Total Epochs: {len(obj_history)}")
print(f"Runtime: {runtime:.6f} s")
print(f"Final Objective Value: {final_obj_saga:e}")
print(f"Non-zero elements in w: {np.sum(np.abs(w_opt) > 1e-4)} / {N}")



print("\n--- Running CVXPY for Ground Truth Verification ---")

w_cvx = cp.Variable(N)
loss = params["b"] - R @ w_cvx

objective = (
    -mu.T @ w_cvx +
    (params["lambda"] / T) * cp.sum(cp.pos(loss)**2) +
    params["alpha_L1"] * cp.norm1(w_cvx)
)

prob = cp.Problem(cp.Minimize(objective))

start_cvx = time.time()
prob.solve(solver=cp.OSQP, verbose=False)
runtime_cvx = time.time() - start_cvx

final_obj_cvx = prob.value



print("\n--- Comparison SAGA vs. CVX ---")
print(f"SAGA Final Obj: {final_obj_saga:e}")
print(f"CVX Final Obj : {final_obj_cvx:e}")

sol_diff_norm = np.linalg.norm(w_opt.reshape(-1) - w_cvx.value)
obj_diff = abs(final_obj_saga - final_obj_cvx)

print(f"Difference in solution norm: {sol_diff_norm:e}")
print(f"Difference in objective:     {obj_diff:e}")

if sol_diff_norm < 1e-3 and obj_diff < 1e-6:
    print("\nSUCCESS: The solution is consistent with CVX.")
else:
    print("\nFAILURE: The solution is significantly different from CVX.")


print("\n--- Final Portfolio Vector (w_opt) ---")
idx = np.where(np.abs(w_opt) > 1e-6)[0]

if len(idx) == 0:
    print("The final w_opt vector is all zeros.")
else:
    print(f"Displaying {len(idx)} non-zero elements:")
    for i in idx:
        print(f"Index {i+1},   w = {w_opt[i].item(): .6e}")



plt.figure(figsize=(8, 5))
plt.plot(obj_history, marker='o', linewidth=2)
plt.xlabel("Epoch", fontsize=14)
plt.ylabel("Objective F(w)", fontsize=14)
plt.title("Prox-SAGA Convergence Curve", fontsize=16)
plt.grid(True, which='both', linestyle="--", linewidth=0.6)
plt.tight_layout()
plt.show()


plt.figure(figsize=(12, 5))
plt.bar(np.arange(N), w_opt.reshape(-1))
plt.xlabel("Asset Index", fontsize=14)
plt.ylabel("Weight Value", fontsize=14)
plt.title("Final Portfolio Weights (w_opt)", fontsize=16)
plt.grid(True, linestyle="--", linewidth=0.5)
plt.tight_layout()
plt.show()
